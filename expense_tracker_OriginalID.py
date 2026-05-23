"""
Bank Statement PDF Parser — SBI
---------------------------------
Accepts a FOLDER containing SBI PDF bank statements.
Outputs:
  1. transaction_summary.xlsx  — Excel report with merchant-wise summary
  2. combined_statements.pdf   — all input PDFs merged into one
 
Usage:
    python parse_bank_statement.py /path/to/folder
    python parse_bank_statement.py /path/to/folder --output my_summary.xlsx --pdf combined.pdf
 
SBI Column order:
    Value Date | Post Date | Details | Ref ID | Debit | Credit | Total Balance
 
Merchant extraction:
    4th slash-delimited segment  e.g. "UPI/DR/612.../THE PERF/..." → "THE PERF"
"""
 
import sys
import re
import argparse
import pdfplumber
import pandas as pd
from pypdf import PdfWriter, PdfReader
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
 
 
# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
 
BANK_KEYWORDS = [
    "sbi", "hdfc", "icici", "axis", "kotak", "yes",
    "pnb", "bob", "canara", "idfc", "indusind", "union", "federal"
]
 
BANK_COLORS = [
    "E8F4FD", "FFF3E0", "F3E5F5", "E8F5E9",
    "FFF8E1", "FCE4EC", "E0F2F1", "EDE7F6",
    "F9FBE7", "E3F2FD"
]
 
# SBI column mapping
COL_MAP = {"date": 0, "details": 2, "ref_id": 3, "debit": 4, "credit": 5, "balance": 6}
 
HEADER_KEYWORDS = [
    "value date", "txn date", "date", "particulars", "narration",
    "withdrawal", "debit", "chq", "ref no"
]
 
 
# ---------------------------------------------------------------------------
# Categorization
# ---------------------------------------------------------------------------
 
CATEGORY_KEYWORDS = {
    "Food & Dining":   ["swiggy", "zomato", "dominos", "pizza", "burger", "food", "restaurant",
                        "cafe", "coffee", "eat", "dining", "hotel", "dhaba", "biryani"],
    "Groceries":       ["bigbasket", "grofers", "blinkit", "zepto", "dunzo", "grocery",
                        "supermarket", "dmart", "reliance fresh", "more store", "vegetables"],
    "Transport":       ["uber", "ola", "rapido", "metro", "irctc", "railway", "bus", "petrol",
                        "fuel", "redbus", "makemytrip", "indigo", "spicejet", "air", "cab"],
    "Shopping":        ["amazon", "flipkart", "myntra", "ajio", "meesho", "nykaa", "snapdeal",
                        "shopify", "retail", "mall", "store", "mart"],
    "Utilities":       ["electricity", "water", "gas", "broadband", "internet", "wifi",
                        "airtel", "jio", "vodafone", "bsnl", "tata sky", "recharge", "bill"],
    "Health":          ["pharmacy", "medical", "hospital", "clinic", "doctor", "apollo",
                        "medplus", "healthkart", "netmeds", "1mg", "medicine"],
    "Entertainment":   ["netflix", "hotstar", "prime", "spotify", "youtube", "gaming",
                        "movie", "pvr", "inox", "bookmyshow", "disney"],
    "Education":       ["udemy", "coursera", "byju", "unacademy", "school", "college",
                        "fees", "tuition", "book", "stationery"],
    "Finance":         ["insurance", "lic", "mutual fund", "sip", "loan", "emi",
                        "credit card", "bank", "investment", "policy"],
}
 
def classify_transaction(merchant: str) -> str:
    merchant_lower = merchant.lower()
    for category, keywords in CATEGORY_KEYWORDS.items():
        if any(kw in merchant_lower for kw in keywords):
            return category
    return "Miscellaneous"
 
 
# ---------------------------------------------------------------------------
# Bank Detection
# ---------------------------------------------------------------------------
 
def detect_bank(filename: str) -> str:
    stem = Path(filename).stem.lower()
    for kw in BANK_KEYWORDS:
        if kw in stem:
            return kw.upper()
    return Path(filename).stem[:20].upper()
 
 
# ---------------------------------------------------------------------------
# Collect PDFs from folder
# ---------------------------------------------------------------------------
 
def collect_pdfs(folder: str) -> list:
    folder_path = Path(folder)
    if not folder_path.is_dir():
        print(f"ERROR: '{folder}' is not a valid folder.")
        sys.exit(1)
    pdfs = sorted(folder_path.glob("*.pdf"))
    if not pdfs:
        print(f"ERROR: No PDF files found in '{folder}'.")
        sys.exit(1)
    print(f"Found {len(pdfs)} PDF(s) in '{folder}':")
    for p in pdfs:
        print(f"  {p.name}  →  [{detect_bank(str(p))}]")
    return [str(p) for p in pdfs]
 
 
# ---------------------------------------------------------------------------
# Merge input PDFs into one combined PDF
# ---------------------------------------------------------------------------
 
def merge_pdfs(pdf_paths: list, output_pdf: str):
    writer = PdfWriter()
    for path in pdf_paths:
        reader = PdfReader(path)
        for page in reader.pages:
            writer.add_page(page)
    with open(output_pdf, "wb") as f:
        writer.write(f)
    print(f"  Combined PDF saved to: {output_pdf}")
 
 
# ---------------------------------------------------------------------------
# PDF Table Extraction
# ---------------------------------------------------------------------------
 
def extract_tables_from_pdf(pdf_path: str) -> list:
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if row and any(cell for cell in row):
                        rows.append(row)
    return rows
 
 
# ---------------------------------------------------------------------------
# Row Parsing
# ---------------------------------------------------------------------------
 
def clean_rows(rows: list) -> pd.DataFrame:
    EXPECTED_COLS = 7
    records = []
    buffer  = None
 
    for row in rows:
        row = [str(c).strip() if c else "" for c in row]
 
        if any(h in row[0].lower() for h in HEADER_KEYWORDS):
            continue
 
        if re.match(r"\d{1,2}[-/]\d{1,2}[-/]\d{2,4}", row[0]):
            if buffer:
                records.append(buffer)
            while len(row) < EXPECTED_COLS:
                row.append("")
            buffer = [
                row[COL_MAP["date"]],
                row[COL_MAP["details"]],
                row[COL_MAP["ref_id"]],
                row[COL_MAP["debit"]],
                row[COL_MAP["credit"]],
                row[COL_MAP["balance"]],
            ]
        elif buffer:
            extra = " ".join(c for c in row if c)
            if extra:
                buffer[1] = (buffer[1] + " " + extra).strip()
 
    if buffer:
        records.append(buffer)
 
    return pd.DataFrame(records, columns=[
        "value_date", "details", "ref_id", "debit", "credit", "balance"
    ])
 
 
# ---------------------------------------------------------------------------
# Merchant Extraction
# ---------------------------------------------------------------------------
 
def extract_merchant(details: str) -> str:
    parts = details.split("/")
    if len(parts) >= 4:
        name = re.sub(r"\s+\d{6,}.*$", "", parts[3]).strip()
        if name:
            return name
    return "UNKNOWN"
 
 
# ---------------------------------------------------------------------------
# Amount Parsing
# ---------------------------------------------------------------------------
 
def parse_amount(val: str) -> float:
    val = re.sub(r"[^\d.]", "", str(val))
    try:
        return float(val) if val else 0.0
    except ValueError:
        return 0.0
 
 
# ---------------------------------------------------------------------------
# Per-PDF Processing
# ---------------------------------------------------------------------------
 
def process_pdf(pdf_path: str, debug: bool = False) -> pd.DataFrame:
    bank_name = detect_bank(pdf_path)
    print(f"  [{bank_name}] Reading: {Path(pdf_path).name}")
 
    raw_data = extract_tables_from_pdf(pdf_path)
    if not raw_data:
        print(f"  [{bank_name}] WARNING: No table data found — skipping.")
        return pd.DataFrame()
 
    if debug:
        print(f"\n  DEBUG — First 5 raw rows from '{Path(pdf_path).name}':")
        for i, row in enumerate(raw_data[:5]):
            print(f"    {i}: {row}")
        print()
 
    df = clean_rows(raw_data)
    if df.empty:
        print(f"  [{bank_name}] WARNING: No transactions parsed — skipping.")
        return pd.DataFrame()
 
    df["bank"]       = bank_name
    df["value_date"] = pd.to_datetime(df["value_date"], dayfirst=True, errors="coerce", format="mixed")
    df["month_name"] = df["value_date"].dt.strftime("%B %Y")
    df["debit_amt"]  = df["debit"].apply(parse_amount)
    df["credit_amt"] = df["credit"].apply(parse_amount)
    df["txn_type"]   = df.apply(
        lambda r: "Credit" if r["credit_amt"] > 0 else ("Debit" if r["debit_amt"] > 0 else "Unknown"),
        axis=1
    )
    df["amount"]   = df.apply(
        lambda r: r["credit_amt"] if r["txn_type"] == "Credit" else r["debit_amt"], axis=1
    )
    df["merchant"] = df["details"].apply(extract_merchant)
    df = df[df["merchant"] != "UNKNOWN"].copy()
 
    print(f"  [{bank_name}] {len(df)} usable transactions found.")
    return df
 
 
# ---------------------------------------------------------------------------
# Summary Builder
# ---------------------------------------------------------------------------
 
def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    summary = (
        df.groupby(["bank", "merchant", "txn_type", "month_name"], sort=False)
        .agg(total_transactions=("amount", "count"), sum_amount=("amount", "sum"))
        .reset_index()
    )
 
    debit_total = (
        df[df["txn_type"] == "Debit"]
        .groupby(["bank", "merchant", "month_name"])["amount"].sum()
        .reset_index().rename(columns={"amount": "total_debit"})
    )
    credit_total = (
        df[df["txn_type"] == "Credit"]
        .groupby(["bank", "merchant", "month_name"])["amount"].sum()
        .reset_index().rename(columns={"amount": "total_credit"})
    )
    net = debit_total.merge(credit_total, on=["bank", "merchant", "month_name"], how="left")
    net["total_credit"] = net["total_credit"].fillna(0)
    net["net_spent"]    = (net["total_debit"] - net["total_credit"]).round(2)
 
    avg = (
        df[df["txn_type"] == "Debit"]
        .groupby(["bank", "merchant", "month_name"])["amount"].mean()
        .reset_index().rename(columns={"amount": "avg_spend_per_month"})
    )
 
    summary = summary.merge(avg, on=["bank", "merchant", "month_name"], how="left")
    summary = summary.merge(
        net[["bank", "merchant", "month_name", "net_spent"]],
        on=["bank", "merchant", "month_name"], how="left"
    )
 
    summary = summary[[
        "bank", "merchant", "txn_type", "total_transactions",
        "sum_amount", "month_name", "avg_spend_per_month", "net_spent"
    ]]
    summary.columns = [
        "Bank", "Merchant Name", "Transaction Type", "Total Transactions",
        "Sum of Amount (Rs)", "Month", "Avg Spend on Merchant (Rs)", "Net Spent (Rs)"
    ]
 
    for col in ["Sum of Amount (Rs)", "Avg Spend on Merchant (Rs)", "Net Spent (Rs)"]:
        summary[col] = summary[col].round(2)
 
    return summary
 
 
# ---------------------------------------------------------------------------
# Amount size buckets
# ---------------------------------------------------------------------------
 
def size_category(amount: float) -> str:
    if amount <= 50:   return "Mini (<=50)"
    if amount <= 200:  return "Small (<=200)"
    if amount <= 500:  return "Medium (<=500)"
    return "Large (>500)"
 
 
# ---------------------------------------------------------------------------
# Analytics Tab Builder
# ---------------------------------------------------------------------------
 
def write_analytics_tab(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Analytics")
 
    title_font   = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    header_font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    body_font    = Font(name="Arial", size=10)
    bold_font    = Font(name="Arial", bold=True, size=10)
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left",   vertical="center")
    currency_fmt = '#,##0.00'
 
    title_fill   = PatternFill("solid", start_color="1F4E79")
    header_fill  = PatternFill("solid", start_color="2E75B6")
    section_fill = PatternFill("solid", start_color="D6E4F0")
    alt_fill     = PatternFill("solid", start_color="EBF3FB")
 
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
 
    debit_df  = df[df["txn_type"] == "Debit"]
    credit_df = df[df["txn_type"] == "Credit"]
 
    row = 1
 
    def write_title(text, row):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font      = title_font
        cell.fill      = title_fill
        cell.alignment = left
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 22
        return row + 1
 
    def write_header(cols, row):
        for c, val in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
        ws.row_dimensions[row].height = 18
        return row + 1
 
    def write_data_row(vals, row, alternate=False):
        fill = alt_fill if alternate else PatternFill()
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font      = body_font
            cell.alignment = center
            cell.fill      = fill
            if isinstance(val, float):
                cell.number_format = currency_fmt
        return row + 1
 
    def write_kv(label, value, row, is_currency=False):
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        lc.font = bold_font;  lc.alignment = left;   lc.fill = section_fill
        vc.font = body_font;  vc.alignment = center
        if is_currency:
            vc.number_format = currency_fmt
        ws.row_dimensions[row].height = 16
        return row + 1
 
    row = write_title("1. Overall Transaction Summary", row)
    row = write_kv("Total Transactions",        len(df),                    row)
    row = write_kv("Total Amount (Rs)",         round(df["amount"].sum(),2), row, True)
    row += 1
 
    row = write_title("2. Credit vs Debit Breakdown", row)
    row = write_header(["Type", "No. of Transactions", "Total Amount (Rs)"], row)
    row = write_data_row(["Debit",  len(debit_df),  round(debit_df["amount"].sum(),  2)], row)
    row = write_data_row(["Credit", len(credit_df), round(credit_df["amount"].sum(), 2)], row, True)
    row += 1
 
    row = write_title("3. Week-wise Spend (Debit Only)", row)
    row = write_header(["Week Starting", "No. of Transactions", "Total Spend (Rs)", "Avg per Transaction (Rs)"], row)
 
    debit_df2 = debit_df.copy()
    debit_df2["week_start"] = debit_df2["value_date"].dt.to_period("W").apply(lambda p: p.start_time.date())
    weekly = (
        debit_df2.groupby("week_start")["amount"]
        .agg(count="count", total="sum")
        .reset_index()
        .sort_values("week_start")
    )
    weekly["avg"] = (weekly["total"] / weekly["count"]).round(2)
 
    for i, (_, wrow) in enumerate(weekly.iterrows()):
        row = write_data_row([
            str(wrow["week_start"]),
            int(wrow["count"]),
            round(wrow["total"], 2),
            round(wrow["avg"],   2)
        ], row, i % 2 == 1)
    row += 1
 
    row = write_title("4. Transaction Size Breakdown", row)
    row = write_header(["Category", "No. of Transactions", "Total Amount (Rs)"], row)
 
    df2 = df.copy()
    df2["size_cat"] = df2["amount"].apply(size_category)
    size_order = ["Mini (<=50)", "Small (<=200)", "Medium (<=500)", "Large (>500)"]
    size_grp = df2.groupby("size_cat")["amount"].agg(count="count", total="sum").reindex(size_order).fillna(0)
 
    for i, (cat, srow) in enumerate(size_grp.iterrows()):
        row = write_data_row([cat, int(srow["count"]), round(srow["total"], 2)], row, i % 2 == 1)
    row += 1
 
    row = write_title("5. Transaction Category Classification (Debit Only)", row)
    row = write_header(["Category", "No. of Transactions", "Total Amount (Rs)", "Avg per Transaction (Rs)"], row)
 
    debit_df3 = debit_df.copy()
    debit_df3["category"] = debit_df3["merchant"].apply(classify_transaction)
    cat_grp = (
        debit_df3.groupby("category")["amount"]
        .agg(count="count", total="sum")
        .reset_index()
        .sort_values("total", ascending=False)
    )
    cat_grp["avg"] = (cat_grp["total"] / cat_grp["count"]).round(2)
 
    for i, (_, crow) in enumerate(cat_grp.iterrows()):
        row = write_data_row([
            crow["category"],
            int(crow["count"]),
            round(crow["total"], 2),
            round(crow["avg"],   2)
        ], row, i % 2 == 1)
 
    return wb
 
 
# ---------------------------------------------------------------------------
# Category Details Tab Builder  ← NEW
# ---------------------------------------------------------------------------
 
CATEGORY_COLORS = {
    "Food & Dining":   "FFF3E0",
    "Groceries":       "E8F5E9",
    "Transport":       "E3F2FD",
    "Shopping":        "FCE4EC",
    "Utilities":       "F3E5F5",
    "Health":          "E0F7FA",
    "Entertainment":   "FFF9C4",
    "Education":       "F1F8E9",
    "Finance":         "EDE7F6",
    "Miscellaneous":   "F5F5F5",
}
 
def write_category_details_tab(wb, df: pd.DataFrame):
    """
    Writes a 'Category Details' sheet listing every transaction row
    alongside its assigned category, sorted by Category → Date.
    """
    ws = wb.create_sheet("Category Details")
 
    header_font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    body_font    = Font(name="Arial", size=10)
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left",   vertical="center")
    header_fill  = PatternFill("solid", start_color="1F4E79")
    currency_fmt = '#,##0.00'
    date_fmt     = 'DD-MMM-YYYY'
 
    col_widths = [18, 22, 45, 25, 15, 15, 18, 20]
    headers    = ["Date", "Bank", "Details", "Merchant", "Type", "Amount (Rs)", "Month", "Category"]
    for i, (w, h) in enumerate(zip(col_widths, headers), start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
        cell = ws.cell(row=1, column=i, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"
 
    detail_df = df.copy()
    detail_df["category"] = detail_df["merchant"].apply(classify_transaction)
    detail_df = detail_df.sort_values(["category", "value_date"]).reset_index(drop=True)
 
    for i, r in detail_df.iterrows():
        cat   = r["category"]
        color = CATEGORY_COLORS.get(cat, "FFFFFF")
        row_fill = PatternFill("solid", start_color=color)
        excel_row = i + 2
 
        values = [
            r["value_date"].date() if pd.notna(r["value_date"]) else "",
            r["bank"],
            r["details"],
            r["merchant"],
            r["txn_type"],
            round(r["amount"], 2),
            r["month_name"],
            cat,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.font      = body_font
            cell.fill      = row_fill
            cell.alignment = left if col in (3, 4) else center
        # Format date and amount columns
        ws.cell(row=excel_row, column=1).number_format = date_fmt
        ws.cell(row=excel_row, column=6).number_format = currency_fmt
 
    # Auto-filter on header row
    ws.auto_filter.ref = f"A1:H{len(detail_df) + 1}"
 
    return wb
 
 
# ---------------------------------------------------------------------------
# Excel Writer
# ---------------------------------------------------------------------------
 
def write_excel(summary: pd.DataFrame, df: pd.DataFrame, output_path: str):
    summary.to_excel(output_path, index=False, sheet_name="Transaction Summary")
 
    wb = load_workbook(output_path)
    ws = wb.active
 
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    body_font   = Font(name="Arial", size=10)
    center      = Alignment(horizontal="center", vertical="center")
 
    col_widths = [14, 30, 20, 22, 22, 18, 28, 22]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width
 
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
 
    banks = summary["Bank"].unique().tolist()
    bank_color_map = {b: BANK_COLORS[i % len(BANK_COLORS)] for i, b in enumerate(banks)}
 
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bank_val = ws.cell(row=i, column=1).value
        row_fill = PatternFill("solid", start_color=bank_color_map.get(bank_val, "FFFFFF"))
        for cell in row:
            cell.font      = body_font
            cell.alignment = center
            cell.fill      = row_fill
 
    currency_fmt = '#,##0.00'
    for col in [5, 7, 8]:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.number_format = currency_fmt
 
    green_font = Font(name="Arial", size=10, color="006100", bold=True)
    red_font   = Font(name="Arial", size=10, color="9C0006", bold=True)
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            if cell.value is not None:
                cell.font = green_font if cell.value >= 0 else red_font
 
    bold_font = Font(name="Arial", size=10, bold=True)
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold_font
 
    # Analytics tab
    wb = write_analytics_tab(wb, df)
 
    # Category Details tab  ← NEW
    wb = write_category_details_tab(wb, df)
 
    # NOTE: Bank Legend sheet intentionally removed
 
    wb.save(output_path)
 
 
# ---------------------------------------------------------------------------
# Entry Point
# ---------------------------------------------------------------------------
 
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Parse SBI bank statement PDFs in a folder → Excel summary + merged PDF."
    )
    parser.add_argument("folder", help="Folder containing PDF bank statements")
    parser.add_argument("--output", default="transaction_summary.xlsx", help="Output Excel filename")
    parser.add_argument("--pdf",    default="combined_statements.pdf",  help="Output merged PDF filename")
    parser.add_argument("--debug",  action="store_true", help="Print raw rows for troubleshooting")
    args = parser.parse_args()
 
    print(f"\nScanning folder: {args.folder}\n")
    pdf_paths = collect_pdfs(args.folder)
 
    print(f"\nMerging PDFs...")
    merge_pdfs(pdf_paths, args.pdf)
 
    print(f"\nExtracting transactions...\n")
    all_dfs = []
    for pdf_path in pdf_paths:
        df = process_pdf(pdf_path, debug=args.debug)
        if not df.empty:
            all_dfs.append(df)
 
    if not all_dfs:
        print("\nERROR: No usable data extracted from any PDF.")
        sys.exit(1)
 
    merged = pd.concat(all_dfs, ignore_index=True)
    print(f"\nTotal transactions across all banks: {len(merged)}")
 
    summary = build_summary(merged)
    write_excel(summary, merged, args.output)
 
    print(f"\n{'='*50}")
    print(f"Done!")
    print(f"  Excel report  : {args.output}")
    print(f"  Combined PDF  : {args.pdf}")
    print(f"  Summary rows  : {len(summary)}")
    print(f"  Banks found   : {', '.join(summary['Bank'].unique())}")
    print(f"{'='*50}")